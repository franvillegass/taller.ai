from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def formatear_excel(path, estilo):
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor=estilo.get("header_color", "2F4F7F"))
    header_font = Font(bold=True, color=estilo.get("font_color_header", "FFFFFF"), size=estilo.get("font_size", 11))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    fill_alt = PatternFill("solid", fgColor=estilo.get("row_alt_color", "F2F2F2"))
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            if i % 2 == 0:
                cell.fill = fill_alt
            cell.font = Font(size=estilo.get("font_size", 11))
            cell.alignment = Alignment(horizontal="left", wrap_text=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save(path)

            


def aplicar_formulas(ws):
    headers = [cell.value for cell in ws[1]]

    def buscar_col(keywords):
        for i, h in enumerate(headers):
            if h and any(k.lower() in h.lower() for k in keywords):
                return i + 1
        return None

    col_total = buscar_col(["total", "subtotal"])
    col_precio = buscar_col(["precio", "price", "unitario"])
    col_cantidad = buscar_col(["cantidad", "quantity", "vendida"])
    col_promedio = buscar_col(["promedio", "average"])

    ultima_fila = ws.max_row

    # Fórmula Total = Precio * Cantidad por fila
    for row in range(2, ultima_fila + 1):
        if col_total and col_precio and col_cantidad:
            precio_cell = ws.cell(row=row, column=col_precio).coordinate
            cantidad_cell = ws.cell(row=row, column=col_cantidad).coordinate
            ws.cell(row=row, column=col_total).value = f"={precio_cell}*{cantidad_cell}"

    # Fila de totales al final
    fila_suma = ultima_fila + 2  # deja una fila vacía de separación

    if col_total:
        col_total_letra = ws.cell(row=2, column=col_total).column_letter
        ws.cell(row=fila_suma, column=col_total).value = f"=SUM({col_total_letra}2:{col_total_letra}{ultima_fila})"
        ws.cell(row=fila_suma, column=col_total - 1).value = "TOTAL"

    if col_cantidad:
        col_cantidad_letra = ws.cell(row=2, column=col_cantidad).column_letter
        ws.cell(row=fila_suma, column=col_cantidad).value = f"=SUM({col_cantidad_letra}2:{col_cantidad_letra}{ultima_fila})"

    if col_promedio:
        col_promedio_letra = ws.cell(row=2, column=col_promedio).column_letter
        ws.cell(row=fila_suma, column=col_promedio).value = f"=AVERAGE({col_promedio_letra}2:{col_promedio_letra}{ultima_fila})"
        ws.cell(row=fila_suma, column=col_promedio - 1).value = "PROMEDIO"

    # Estilo de la fila de totales
    from openpyxl.styles import Font
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=fila_suma, column=col)
        cell.font = Font(bold=True, size=11)

